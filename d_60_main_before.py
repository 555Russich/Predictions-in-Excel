import time
import pandas as pd
import pandas_datareader as web
import datetime as dt
import numpy as np
from sklearn.preprocessing import MinMaxScaler
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense, Dropout, LSTM
from openpyxl import load_workbook


def setup():
    filename_excel = 'predictions.xlsx'
    return filename_excel


def get_data_from_excel(filename_excel):
    data_excel = pd.read_excel(filename_excel)
    wb = load_workbook(filename_excel)
    return data_excel, wb


def sort_data_excel(data_excel):
    list_tickets = data_excel['Ticket'].values
    # list_price_before = data_excel['Price close'].values
    # list_price_prediction = data_excel['Price close prediction'].values
    # list_price_after = data_excel['Price 17.05'].values
    # list_percents_prediction = data_excel['Prediction of delta percents'].values
    # list_percents_after = data_excel['%(17 - predict)'].values
    # list_result_of_predictions = data_excel['Result'].values
    return list_tickets


def timer_start():
    start_tt = time.time()
    return start_tt


def timer_tt(start):
    tt = time.time() - start
    print(f'tt:{round(tt,2)} sec')
    return tt


def prediction_process(list_tickets):
    # Load Data
    list_of_predictions = []
    list_price_before = []
    prediction_days = 60
    number_of_prediction = 3
    counter_company = 0
    for company in list_tickets:

        start = dt.datetime.now() - dt.timedelta(weeks=156)
        end = dt.datetime.now()

        data = web.DataReader(company, 'yahoo', start, end)

        # Get close price of last day
        price_before = data['Close'].values[-1]
        list_price_before.append(price_before)

        # Prepare Data
        scaler = MinMaxScaler(feature_range=(0, 1))
        scaled_data = scaler.fit_transform(data['Close'].values.reshape(-1, 1))

        x_train = []
        y_train = []

        for x in range(prediction_days, len(scaled_data)):
            x_train.append(scaled_data[x - prediction_days:x, 0])
            y_train.append(scaled_data[x, 0])

        x_train, y_train = np.array(x_train), np.array(y_train)
        x_train = np.reshape(x_train, (x_train.shape[0], x_train.shape[1], 1))

        # Build The Model
        prediction_list_to_average = []
        for n in range(number_of_prediction):
            model = Sequential()

            model.add(LSTM(units=50, return_sequences=True, input_shape=(x_train.shape[1], 1)))
            model.add(Dropout(0.2))
            model.add(LSTM(units=50, return_sequences=True))
            model.add(Dropout(0.2))
            model.add(LSTM(units=50))
            model.add(Dropout(0.2))
            model.add(Dense(units=1))  # Prediction of the next closing value

            model.compile(optimizer='adam', loss='mean_squared_error')
            model.fit(x_train, y_train, epochs=25, batch_size=32)

            ''' Test the model Accuracy on Existing Data '''

            # Load test data

            test_start = dt.datetime.now() - dt.timedelta(weeks=156)
            test_end = dt.datetime.now()

            test_data = web.DataReader(company, 'yahoo', test_start, test_end)
            actual_prices = test_data['Close'].values

            total_dataset = pd.concat((data['Close'], test_data['Close']), axis=0)

            model_inputs = total_dataset[len(total_dataset) - len(test_data) - prediction_days:].values
            model_inputs = model_inputs.reshape(-1, 1)
            model_inputs = scaler.transform(model_inputs)

            # Make predictions on test data

            x_test = []

            for x in range(prediction_days, len(model_inputs)):
                x_test.append(model_inputs[x - prediction_days:x, 0])

            x_test = np.array(x_test)
            x_test = np.reshape(x_test, (x_test.shape[0], x_test.shape[1], 1))

            predicted_prices = model.predict(x_test)
            predicted_prices = scaler.inverse_transform(predicted_prices)

            # Predict next day
            real_data = [model_inputs[len(model_inputs) + 1 - prediction_days:len(model_inputs + 1), 0]]
            real_data = np.array(real_data)
            real_data = np.reshape(real_data, (real_data.shape[0], real_data.shape[1], 1))

            prediction = model.predict(real_data)
            prediction = scaler.inverse_transform(prediction)
            prediction_list_to_average.append(prediction[0][0])
        prediction_average = sum(prediction_list_to_average) / len(prediction_list_to_average)
        list_of_predictions.append(round(prediction_average, 2))
        print(f'№:{counter_company}\ncompany:{company}\nprediction price:{prediction_average}')
        counter_company += 1
    return list_of_predictions, list_price_before


def getting_percentage(list_price_before, list_of_predictions):
    i = 0
    list_percentage = []
    while i < len(list_price_before):
        percent = (list_of_predictions[i] - list_price_before[i]) / list_of_predictions[i]
        percent = '{0:+.2%}'.format(percent)
        list_percentage.append(percent)
        i += 1
    return list_percentage


def create_data_frame(list_tickets, list_price_before, list_of_predictions, list_percentage):
    data = {
        'Ticket': list_tickets,
        'Price before': list_price_before,
        'Price prediction': list_of_predictions,
        'Percentage 1': list_percentage
    }

    data_excel_new = pd.DataFrame.from_dict(data, orient='index')  # To make data frame with different length of array
    data_excel_new = data_excel_new.transpose()
    return data_excel_new


def new_sheet(filename_excel, wb, data_excel_new):
    list_sheetnames = wb.sheetnames
    new_sheetname = list_sheetnames[-1][:-1] + str(int(list_sheetnames[-1][-1]) + 1)
    wb.create_sheet(new_sheetname)
    with pd.ExcelWriter(filename_excel, engine='openpyxl', mode='a') as writer:
        data_excel_new.to_excel(writer, new_sheetname)


if __name__ == '__main__':
    start_tt = timer_start()
    filename_excel = setup()
    data_excel, wb = get_data_from_excel(filename_excel)

    ''' Put excel data and take needed lists from there '''
    list_tickets = sort_data_excel(data_excel)

    '''Process of prediction. Days to look back to predict another = 60.
     Predict price 3 times then find average of them '''
    list_of_predictions, list_price_before = prediction_process(list_tickets)

    ''' Calculate percentage between prediction and price before '''
    list_percentage = getting_percentage(list_price_before, list_of_predictions)

    ''' Create panda data frame to write it in excel '''
    data_excel_new = create_data_frame(list_tickets, list_price_before, list_of_predictions, list_percentage)

    ''' Append new sheet which name is last sheet + 1 (if last 'Sheet1', append 'Sheet2') 
    in excel file and push created data frame '''
    new_sheet(filename_excel, wb, data_excel_new)

    tt = timer_tt(start_tt)